import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell import MergedCell
from openpyxl.styles import PatternFill
from os import system


def parser_merged_cell(sheet: Worksheet, row, col):
    """
    检查是否为合并单元格并获取对应行列单元格的值。
    如果是合并单元格，则取合并区域左上角单元格的值作为当前单元格的值,否则直接返回该单元格的值
    :param sheet: 当前工作表对象
    :param row: 需要获取的单元格所在行
    :param col: 需要获取的单元格所在列
    :return:
    """
    cell = sheet.cell(row=row, column=col)
    if isinstance(cell, MergedCell):  # 判断该单元格是否为合并单元格
        for merged_range in sheet.merged_cells.ranges:  # 循环查找该单元格所属的合并区域
            if cell.coordinate in merged_range:
                # 获取合并区域左上角的单元格作为该单元格的值返回
                cell = sheet.cell(row=merged_range.min_row, column=merged_range.min_col)
                break
    return cell


def check_modification(gd_sheet, tj_sheet):
    """
    检查表格格式是否有改动
    :return:
    """
    if gd_sheet['D1'].value != '业务类型':
        raise ValueError('工单表D1不是业务类型')
    if gd_sheet['F1'].value != '位置':
        raise ValueError('工单表F1不是位置')
    if gd_sheet['P1'].value != '创建人':
        raise ValueError('工单表P1不是创建人')

    if tj_sheet['C1'].value != '片区自报':
        raise ValueError('统计表C1不是片区自报')
    if tj_sheet['D1'].value != '北岸巡查':
        raise ValueError('统计表D1不是北岸巡查')
    if tj_sheet['E1'].value != '合计':
        raise ValueError('统计表E1不是合计')
    if tj_sheet['A20'].value != '六河七园':
        raise ValueError('统计表A20不是六河七园')
    return


def count_initial():
    """
    count initial
    :return:
    """
    lvbaoshi = {"绿化": {}, "保洁": {}, "市政（道桥）": {}, "市政（排水）": {}, "六河七园": {}}
    color = {"绿化": {}, "保洁": {}, "市政（道桥）": {}, "市政（排水）": {}, "六河七园": {}}
    for key in lvbaoshi:
        if key == "市政（道桥）" or key == "市政（排水）":
            lvbaoshi[key]['西片区'] = {}
            lvbaoshi[key]['东片区'] = {}
            lvbaoshi[key]['南片区'] = {}
            lvbaoshi[key]['未上报片区'] = {}
        elif key == "绿化" or key == "保洁":
            lvbaoshi[key]["青山绿水"] = {}
            lvbaoshi[key]["恒生"] = {}
            lvbaoshi[key]["黄岛园林"] = {}
            lvbaoshi[key]["未上报片区"] = {}
    lvbaoshi["市政（排水）"]['东北片区'] = {}
    lvbaoshi["市政（排水）"]['东南片区'] = {}
    lvbaoshi["六河七园"]['白沙河运动公园'] = {}
    lvbaoshi["六河七园"]['墨水河'] = {}
    lvbaoshi["六河七园"]['五水绕城'] = {}
    lvbaoshi["六河七园"]['七园'] = {}
    # lvbaoshi["六河七园"]['未上报片区'] = {}
    for key1 in lvbaoshi:
        for key2 in lvbaoshi[key1]:
            lvbaoshi[key1][key2]['片区自报'] = 0
            lvbaoshi[key1][key2]['北岸巡查'] = 0
    weishangbaohanghao = {"市政（道桥）": [], "市政（排水）": [], "绿化": [], "保洁": []}
    beian_name = ['江坤', '纪玉超', '王锡鹏', '栾延森', 'us系统账号']
    color['绿化']['青山绿水'] = 'FFB7DD'
    color['绿化']['恒生'] = 'FF0088'
    color['绿化']['黄岛园林'] = '8C0044'
    color['保洁']['青山绿水'] = 'FFDDAA'
    color['保洁']['恒生'] = 'FF8800'
    color['保洁']['黄岛园林'] = 'BB5500'
    color['市政（道桥）'] = 'EEFFBB'
    color['市政（排水）'] = 'BBFF00'
    return lvbaoshi, beian_name, weishangbaohanghao, color


def is_liuheqiyuan(yewu_cell, lvbaoshi, beianORpianqu):
    liuheqiyuan = [["白沙河运动公园"],
                   ["墨水河"],
                   ["五水绕城", "虹子河", "小北曲河", "南疃河", "爱民河"],
                   ["七园", "雕塑文化园", "新天地", "国学公园", "人民广场", "行政广场", "北中轴", "民生公园"]]
    for category in liuheqiyuan:
        for i in category:
            if i in gd_sheet['F' + str(yewu_cell.row)].value:
                lvbaoshi["六河七园"][category[0]][beianORpianqu] += 1
                return True
    return False


def go_count(gd_sheet, lvbaoshi, beian_name, weishangbaohanghao, color):
    """
    统计工单表里得数据，记录在lvbaoshi里
    :param gd_sheet:
    :param lvbaoshi:
    :return:
    """
    yewu_cells = gd_sheet['D']

    for yewu_cell in yewu_cells:
        if yewu_cell.value == None:
            break
        yewuleixing = None
        if yewu_cell.value == '环卫' or yewu_cell.value == '环卫保洁':
            yewuleixing = "保洁"
        elif yewu_cell.value == '园林绿化':
            yewuleixing = "绿化"
        elif yewu_cell.value == '市政设施':
            yewuleixing = "市政"
        # else:
        # print('业务类型不对')

        chuangjianren_cell = gd_sheet['P' + str(yewu_cell.row)]
        if chuangjianren_cell.value in beian_name or "us系统账号" in chuangjianren_cell.value:
            beianORpianqu = "北岸巡查" 
        else :
            beianORpianqu ="片区自报"

        # 六河七园
        if is_liuheqiyuan(yewu_cell, lvbaoshi, beianORpianqu):
            continue

        # 保洁、绿化
        if yewuleixing == "保洁" or yewuleixing == "绿化":

            weizhi_cell = gd_sheet['F' + str(yewu_cell.row)]

            if "青山绿水" in weizhi_cell.value:
                lvbaoshi[yewuleixing]["青山绿水"][beianORpianqu] += 1
                weizhi_cell.fill = PatternFill('solid', fgColor=color[yewuleixing]["青山绿水"])
            elif "恒生" in weizhi_cell.value:
                lvbaoshi[yewuleixing]["恒生"][beianORpianqu] += 1
                weizhi_cell.fill = PatternFill('solid', fgColor=color[yewuleixing]["恒生"])
            elif "黄岛园林" in weizhi_cell.value:
                lvbaoshi[yewuleixing]["黄岛园林"][beianORpianqu] += 1
                weizhi_cell.fill = PatternFill('solid', fgColor=color[yewuleixing]["黄岛园林"])
            else:
                lvbaoshi[yewuleixing]["未上报片区"][beianORpianqu] += 1
                weishangbaohanghao[yewuleixing].append(weizhi_cell.row)

        # 市政
        if yewuleixing == "市政":

            weizhi_cell = gd_sheet['F' + str(yewu_cell.row)]
            daoqiaoORpaishui = None
            if "市政道桥" in weizhi_cell.value:
                daoqiaoORpaishui = "市政（道桥）"
            elif "道桥排水" in weizhi_cell.value or "市政排水" in weizhi_cell.value:
                daoqiaoORpaishui = "市政（排水）"

            weizhi_cell.fill = PatternFill('solid', fgColor=color[daoqiaoORpaishui])
            if "西部片区" in weizhi_cell.value:
                lvbaoshi[daoqiaoORpaishui]['西片区'][beianORpianqu] += 1
            elif "东部片区" in weizhi_cell.value:
                lvbaoshi[daoqiaoORpaishui]['东片区'][beianORpianqu] += 1
            elif "南部片区" in weizhi_cell.value:
                lvbaoshi[daoqiaoORpaishui]['南片区'][beianORpianqu] += 1
            elif "东南" in weizhi_cell.value:  # 有风险
                lvbaoshi[daoqiaoORpaishui]['东南片区'][beianORpianqu] += 1
            elif "东北" in weizhi_cell.value:
                lvbaoshi[daoqiaoORpaishui]['东北片区'][beianORpianqu] += 1
            else:
                lvbaoshi[daoqiaoORpaishui]['未上报片区'][beianORpianqu] += 1
                weishangbaohanghao[daoqiaoORpaishui].append(weizhi_cell.row)
    print(lvbaoshi, '\n')
    return lvbaoshi, weishangbaohanghao


def write_xlsx(tj_sheet, lvbaoshi, tj_workbook, weishangbaohanghao, tongji_path):
    """
    统计结果写道xlsx里
    :param tj_sheet:
    :param lvbaoshi:
    :param tj_workbook:
    :return:
    """
    pianqu_cells = tj_sheet['B']
    for pianqu_cell in pianqu_cells:
        yewu = tj_sheet['A' + str(pianqu_cell.row)]
        yewu = parser_merged_cell(tj_sheet, yewu.row, yewu.column)
        if yewu.value in ["绿化", "保洁", "市政（道桥）", "市政（排水）"]:
            tj_sheet['C' + str(pianqu_cell.row)].value = lvbaoshi[yewu.value][pianqu_cell.value]["片区自报"]
            tj_sheet['D' + str(pianqu_cell.row)].value = lvbaoshi[yewu.value][pianqu_cell.value]["北岸巡查"]
            tj_sheet['E' + str(pianqu_cell.row)].value = lvbaoshi[yewu.value][pianqu_cell.value]["片区自报"] + \
                                                         lvbaoshi[yewu.value][pianqu_cell.value]["北岸巡查"]
        # if yewu.value == "六河七园":

    tj_sheet['C20'].value = lvbaoshi['六河七园']['白沙河运动公园']['片区自报']
    tj_sheet['D20'].value = lvbaoshi['六河七园']['白沙河运动公园']['北岸巡查']
    tj_sheet['E20'].value = tj_sheet['C20'].value + tj_sheet['D20'].value
    tj_sheet['C21'].value = lvbaoshi['六河七园']['墨水河']['片区自报']
    tj_sheet['D21'].value = lvbaoshi['六河七园']['墨水河']['北岸巡查']
    tj_sheet['E21'].value = tj_sheet['C21'].value + tj_sheet['D21'].value
    tj_sheet['C22'].value = lvbaoshi['六河七园']['五水绕城']['片区自报']
    tj_sheet['D22'].value = lvbaoshi['六河七园']['五水绕城']['北岸巡查']
    tj_sheet['E22'].value = tj_sheet['C22'].value + tj_sheet['D22'].value
    tj_sheet['C23'].value = lvbaoshi['六河七园']['七园']['片区自报']
    tj_sheet['D23'].value = lvbaoshi['六河七园']['七园']['北岸巡查']
    tj_sheet['E23'].value = tj_sheet['C23'].value + tj_sheet['D23'].value
    tj_sheet['C25'].value = sum([lvbaoshi[key1][key2]["片区自报"] for key1 in lvbaoshi.keys() for key2 in lvbaoshi[key1]])
    tj_sheet['D25'].value = sum([lvbaoshi[key1][key2]["北岸巡查"] for key1 in lvbaoshi.keys() for key2 in lvbaoshi[key1]])
    tj_sheet['E25'].value = tj_sheet['C25'].value + tj_sheet['D25'].value

    tj_workbook.save(tongji_path)

    print(f'市政（道桥）未上报片区：{lvbaoshi["市政（道桥）"]["未上报片区"]}, 行号：{weishangbaohanghao["市政（道桥）"]}\n'
          f'市政（排水）未上报片区：{lvbaoshi["市政（排水）"]["未上报片区"]}, 行号：{weishangbaohanghao["市政（排水）"]}\n'
          f'绿化未上报片区：{lvbaoshi["绿化"]["未上报片区"]}, 行号：{weishangbaohanghao["绿化"]}\n'
          f'保洁未上报片区：{lvbaoshi["保洁"]["未上报片区"]}, 行号：{weishangbaohanghao["保洁"]}')
    system('pause')
    return


if __name__ == '__main__':
    # read from xlsx
    gongdan_path = '工单报表.xlsx'
    tongji_path = '统计.xlsx'
    gd_workbook, gd_sheet, tj_workbook, tj_sheet = None, None, None, None
    if input("是否修改默认excel路径, 输入 y 修改, 其他字符不修改\n").strip() == 'y':
        gongdan_path = input("输入工单报表路径，例如: G:\工单报表.xlsx\n").strip()
        tongji_path = input("输入统计表路径，例如: G:\统计.xlsx\n").strip()

    # 工单
    try:
        gd_workbook = openpyxl.load_workbook(gongdan_path)
        gd_sheet = gd_workbook['工单列表']
        # 统计
        tj_workbook = openpyxl.load_workbook(tongji_path)
        tj_sheet = tj_workbook['Sheet1']
    except:
        print("当前目录没有 工单报表.xlsx 和 统计.xlsx")
        quit()

    # check modification
    check_modification(gd_sheet, tj_sheet)

    # count initial
    lvbaoshi, beian_name, weishangbaohanghao, color = count_initial()
    print("北岸巡查人员默认名单:'江坤', '纪玉超', '王锡鹏', '栾延森', 'us系统账号'")
    mod_xuncha = input("是否修改北岸巡查人员默认名单？输入 y 开始全部修改, a开始追加修改, 其他字符不修改\n").strip()
    if mod_xuncha == 'y' or mod_xuncha == 'a':
        if mod_xuncha == 'y':
            beian_name = []
        while True:
            renming = input("输入人名，一次输入一个，全输完了输入 q\n").strip()
            if renming == 'q':
                break
            beian_name.append(renming)

    # let's go
    go_count(gd_sheet, lvbaoshi, beian_name, weishangbaohanghao, color)

    # write
    write_xlsx(tj_sheet, lvbaoshi, tj_workbook, weishangbaohanghao, tongji_path)
    gd_workbook.save(gongdan_path)

# tj_sheet['C2'].value = lvbaoshi["绿化"]["青山绿水"]["片区自报"]
# tj_sheet['D2'].value = lvbaoshi["绿化"]["青山绿水"]["北岸巡查"]
# tj_sheet['C3'].value = lvbaoshi["绿化"]["恒生"]["片区自报"]
# tj_sheet['D3'].value = lvbaoshi["绿化"]["恒生"]["北岸巡查"]
# tj_sheet['C4'].value = lvbaoshi["绿化"]["黄岛园林"]["片区自报"]
# tj_sheet['D4'].value = lvbaoshi["绿化"]["黄岛园林"]["北岸巡查"]
# tj_sheet['C5'].value = lvbaoshi["绿化"]["未上报片区"]["片区自报"]
# tj_sheet['D5'].value = lvbaoshi["绿化"]["未上报片区"]["北岸巡查"]
